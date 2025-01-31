import os
import sys
import json
import traceback
import pandas as pd
from datetime import datetime
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ✅ Carregar configurações do arquivo JSON
with open("config.json", "r", encoding="utf-8") as config_file:
    config = json.load(config_file)

MONGO_URI = config["mongodb_uri"]
BANCO_DADOS = config["banco_dados"]
COLECAO = config["colecao"]
DIRETORIO_EXPORTACAO = config["caminho_pasta_exportacao"]
CAMINHO_LOGO = config["caminho_logo"]

# ✅ Criar diretório de exportação, se não existir
if not os.path.exists(DIRETORIO_EXPORTACAO):
    os.makedirs(DIRETORIO_EXPORTACAO)

# ✅ Verificar argumento
if len(sys.argv) < 2:
    raise ValueError("O nome do documento (importador) deve ser passado como argumento.")

nome_documento = sys.argv[1]
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
nome_unico = f"{nome_documento}_{timestamp}"
arquivo_excel = os.path.join(DIRETORIO_EXPORTACAO, f"{nome_unico}.xlsx")
arquivo_pdf = os.path.join(DIRETORIO_EXPORTACAO, f"{nome_unico}.pdf")

# ✅ Conectar ao MongoDB
try:
    client = MongoClient(MONGO_URI)
    db = client[BANCO_DADOS]
    collection = db[COLECAO]
except Exception as e:
    print(f"Erro ao conectar ao MongoDB: {e}")
    traceback.print_exc()
    sys.exit(1)

# ✅ Buscar dados do MongoDB
try:
    dados = list(collection.find({"Importador": nome_documento}))
    if not dados:
        raise ValueError(f"Nenhum processo encontrado para o importador: {nome_documento}")
except Exception as e:
    print(f"Erro ao buscar dados no MongoDB: {e}")
    traceback.print_exc()
    sys.exit(1)

# ✅ Converter os dados para um DataFrame e ajustar colunas
df = pd.DataFrame(dados)
campos_desejados = [
    "Ref_USA", "Exportador", "SR", "Produto", "FreeTime", "VencimentoFreeTime", "VencimentoFMA",
    "Navio", "DataDeAtracacao", "DocRecebidos",
    "DataRecebOriginais", "StatusDoProcesso", "DI", "ParametrizacaoDI"
]
df = df[campos_desejados]
df.columns = [
    "Ref. USA", "Exportador", "Ref. Imp", "Produto", "Free Time", "Venc. Free Time", "Venc. FMA",
    "Navio", "Data de Atracação", "Documentos Recebidos",
    "Data de Recebimento Originais", "Status do Processo", "DI", "Parametrização DI"
]

# ✅ Criar Excel e formatar
wb = Workbook()
ws = wb.active
ws.title = "Planilha1"

# Adicionar logo
if os.path.exists(CAMINHO_LOGO):
    img = Image(CAMINHO_LOGO)
    img.width = 200
    img.height = 80
    ws.add_image(img, "A1")

# Adicionar título
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(df.columns) + 1)
doc_name_cell = ws.cell(row=1, column=2, value=nome_documento)
doc_name_cell.font = Font(size=16, bold=True)
doc_name_cell.alignment = Alignment(horizontal="center", vertical="center")

# Adicionar "FollowUp"
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns) + 1)
followup_cell = ws.cell(row=2, column=1, value="FollowUp")
followup_cell.font = Font(size=22, bold=True)
followup_cell.alignment = Alignment(horizontal="center", vertical="center")

# Adicionar cabeçalhos
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
for col_idx, col_name in enumerate(df.columns, start=2):
    cell = ws.cell(row=3, column=col_idx, value=col_name)
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Adicionar dados
for r_idx, row in df.iterrows():
    for c_idx, value in enumerate(row, start=2):
        cell = ws.cell(row=r_idx + 4, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Ajustar largura das colunas
for col_idx in range(2, len(df.columns) + 2):
    column_letter = get_column_letter(col_idx)
    ws.column_dimensions[column_letter].width = 15

# Salvar Excel
wb.save(arquivo_excel)
print(f"Arquivo Excel salvo em: {arquivo_excel}")

# ✅ Converter para PDF usando LibreOffice
def converter_para_pdf(input_file, output_file):
    try:
        comando = f'soffice --headless --convert-to pdf "{input_file}" --outdir "{os.path.dirname(output_file)}"'
        os.system(comando)
        print(f"PDF gerado com sucesso: {output_file}")
    except Exception as e:
        print(f"Erro ao converter para PDF: {e}")

converter_para_pdf(arquivo_excel, arquivo_pdf)
